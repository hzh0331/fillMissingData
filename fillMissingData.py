import datetime

import openpyxl
import xlrd
from xlrd import xldate_as_tuple


def readFile():
    RespWorkBook = xlrd.open_workbook('baby 505/505_2_5624_1.xlsx');
    HRWorkBook = xlrd.open_workbook('baby 505/Baby 505 Numeric with ALL PHASES VALIDATED.xlsx');

    RespContent = RespWorkBook.sheet_by_index(0);  # sheet索引从0开始
    HRContent = HRWorkBook.sheet_by_index(0);

    RespDate = []
    RespData = []
    HRDate = []
    HRData = []

    for i in range(RespContent.nrows):
        if i != 0:
            sCell = RespContent.cell(i, 0)
            if sCell.value == "":
                break
            date = xldate_as_tuple(sCell.value, 0)
            RespDate.append(changeTupleIntoDate(date))
            RespData.append(RespContent.cell(i, 1).value)

    for i in range(HRContent.nrows):
        if i != 0:
            sCell = HRContent.cell(i, 0)
            if sCell.value == "":
                break
            date = xldate_as_tuple(sCell.value, 0)
            HRDate.append(changeTupleIntoDate(date))
            HRData.append(HRContent.cell(i, 3).value)

    # for i in range(len(HRDate)):
    #     if i>=len(HRDate):
    #         break
    #     if HRDate[i] not in RespDate:
    #         HRDate.remove(HRDate[i])
    #         HRData.remove(HRData[i])
    tempCounter = 0
    while True:
        if HRDate[tempCounter] not in RespDate:
            print(HRDate[tempCounter])
            HRDate.remove(HRDate[tempCounter])
            HRData.remove(HRData[tempCounter])
        else:
            tempCounter = tempCounter + 1
        if tempCounter >= len(HRDate):
            break

    for i in range(len(RespDate)):
        if RespDate[i] not in HRDate:
            if i == 0:
                HRDate.insert(0, RespDate[0])
                HRData.insert(0, HRData[0])
            else:
                index = HRDate.index(RespDate[i-1])
                HRDate.insert(index+1, RespDate[i])
                if index == len(HRData) - 1:
                    HRData.insert(index + 1, HRData[index])
                else:
                    HRData.insert(index+1, (HRData[index] + HRData[index+1])/2)
    writeDataIntoFile(RespDate, RespData, HRDate, HRData)

def writeDataIntoFile(RespDate, RespData, HRDate, HRData):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    titleName = ["Date Time", "Resp Data", "", "", "", "Date Time", "HR Data"]
    for i in range(len(titleName)):
        sheet.cell(row=1,column=1+i, value=titleName[i])
    for i in range(len(RespDate)):
        sheet.cell(row=2+i, column=1, value=RespDate[i])
        sheet.cell(row=2+i, column=2, value=RespData[i])
        if i < len(HRDate):
            sheet.cell(row=2+i, column=6, value=HRDate[i])
            sheet.cell(row=2+i, column=7, value=HRData[i])
    workbook.save("baby 505 filled/505_2_5624_1_filled.xlsx")


def changeTupleIntoDate(dateTuple):
    stamp = datetime.datetime(2011, 1, 3, dateTuple[3], dateTuple[4], dateTuple[5])
    return stamp

def addSecond(stamp):
    now = stamp + datetime.timedelta(seconds=1)
    return now

readFile()