import openpyxl
import xlrd
import numpy as np

def readFile():
    workBook = xlrd.open_workbook("test_new/513_hr.xlsx");
    sheet = workBook.sheet_by_index(0);
    SPO2Data = []
    HRData = []
    RRData = []
    for i in range(sheet.nrows):
        SPO2Data.append(sheet.cell(i, 1).value)
        HRData.append(sheet.cell(i,2).value)
        RRData.append(sheet.cell(i,3).value)
    return SPO2Data, HRData, RRData

def calculateAverage(values, begin, end):
    targetValue = values[begin:end]
    return np.mean(targetValue)

def writeExcel(SPO2Average, HRAverage, RRAverage):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    for i in range(5):
        sheet.cell(row=1, column=i+1, value=HRAverage[i])
    for i in range(5):
        sheet.cell(row=1, column=6+i, value=RRAverage[i])
    for i in range(5):
        sheet.cell(row=1, column=11+i, value=SPO2Average[i])
    sheet.cell(row=1, column=16, value="")
    for i in range(3):
        sheet.cell(row=1, column=17+i, value=HRAverage[5+i])
        sheet.cell(row=1, column=17 + i, value=RRAverage[5 + i])
        sheet.cell(row=1, column=17 + i, value=SPO2Average[5 + i])
    workbook.save("output")

SPO2Data, HRData, RRData = readFile()
#phase 1, 2, 3, 4, 5, 3-1, 3-2, 3-3
beginEndPair=[[1,61],[61,121],[121,1462],[1462,1521],[1521,1580],[121, 181], [792,852], 1402]
SPO2Average = []
HRAverage = []
RRAverage = []
for i in beginEndPair:
    SPO2Average.append(calculateAverage(SPO2Data,i[0], i[1]))
    HRAverage.append(calculateAverage(HRData,i[0], i[1]))
    RRAverage.append(calculateAverage(RRData,i[0], i[1]))
writeExcel(SPO2Average, HRAverage, RRAverage)